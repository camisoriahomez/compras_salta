#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import bs4
import requests
import re
from datetime import date
import openpyxl
from openpyxl.styles import Font
  

def armar_excel(compras):
    # creo un nuevo libro vacio
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = 'Reporte'
    # hoy
    hoy = date.today()
    sh.cell(row=1, column=1).value = 'Fecha'
    sh.cell(row=1, column=1).font = Font(bold=True)
    sh.cell(row=1, column=2).value = hoy.strftime("%d/%m/%Y")    
    
    # header
    keys = list(compras[0].keys())
    fila_header = 3
    for i, key in enumerate(keys):
        sh.cell(row=fila_header, column=i+1).value = key
        sh.cell(row=fila_header, column=i+1).font = Font(bold=True)
        
    # datos
    fila_inic = 4
    for c, compra in enumerate(compras):
        dict_actual = compras[c]
        for col, v in enumerate(dict_actual.values()):
            if col == 4:
                sh.cell(row=fila_inic+c, column=col+1).value = '=HYPERLINK("{}", "{}")'.format(v, "link")
            else:
                sh.cell(row=fila_inic+c, column=col+1).value = v       
        
    # ajusto
    sh.column_dimensions['A'].width = 40
    sh.column_dimensions['B'].width = 20
    sh.column_dimensions['C'].width = 12
    sh.column_dimensions['D'].width = 30
    sh.column_dimensions['E'].width = 8
    
    # guardo archivo
    wb.save('Reporte-' + hoy.strftime("%b-%d-%Y") + '.xlsx')


def parsear_cuerpo(soup, web):
    # saco el organismo
    art_org = soup.find_all('div', class_="publicacion-fila")[1]
    org = art_org.get_text().split('\n')[2]
    # link
    art_link = soup.find('div',style="padding:4px 0 0 0;")
    atr_link = art_link.input['onclick'].split("'")[1]
    link = web + atr_link
    return org, link


def parsear_encabezado(soup):
    art_encab = soup.find('div', class_="publicacion-encabezado")
    encab_txt = art_encab.get_text().split('\n')
    encab_txt = [i for i in encab_txt if i]
    # fecha
    regex_fecha = re.compile(r'\d+/\d+/\d+')
    fecha = regex_fecha.search(encab_txt[1]).group()
    # hora
    regex_hora = re.compile(r'\d+:\d+')
    hora = regex_hora.search(encab_txt[2]).group()
    # nro contratacion 
    regex_nro = re.compile(r'\d+/\d+')
    if regex_nro.search(encab_txt[3]):
        nro = regex_nro.search(encab_txt[3]).group()
    else:
        nro = ''
    return fecha, hora, nro
    

def parsear_objeto(soup):
    # busco en el html donde esta definido el objeto
    art_objeto = soup.find('div', class_="publicacion-fila")
    objs = art_objeto.get_text().split('\n')[2] 
    return objs.split() if len(objs.split())>1 else objs   
    

def objeto_matches(soup, keywords):
    words = parsear_objeto(soup)
    for word in words:
        if word.lower() in keywords: return True
    return False
    

def parsear_pagina(soup, lst, keywords, web):
    ''' Toma una web parseada en html como elemento soup y 
    una lista lst donde se guarda la informacion que
    coincide con las palabras en la lista keywords.
    '''
    # busco todos los articulos q contienen la info
    articulos = soup.find_all('article', class_="publicacion")
    # analizo cada uno
    for art in articulos:
        dic_aux = {}
        if objeto_matches(art, keywords):
            # objeto
            lista_objeto = parsear_objeto(art)
            dic_aux['Objeto'] = ' '.join(lista_objeto)
            # del encabezado
            fecha, hora, nro = parsear_encabezado(art)
            dic_aux['Fecha Apertura'] = fecha
            dic_aux['Hora Apertura'] = hora
            #dic_aux['Nro Contratacion'] = nro
            # del cuerpo
            org, link = parsear_cuerpo(art, web)
            dic_aux['Organismo'] = org
            dic_aux['Link'] = link
        if dic_aux:
            lst.append(dic_aux)   


def main(keywords):
    web = 'http://compras.salta.gov.ar/'
    # lista para guardar la info
    compras = []
    for i in range(0,291,5):
        if not i:
            # descargar la pagina
            r = requests.get(web)
        else:
            extra = 'publico/publicacionactual/panelfiltrobusqueda/'
            web_new = web + extra + str(i)
            # descargar la pagina
            r = requests.get(web_new)
        r.raise_for_status()
        # armo el objeto soup
        fullweb_soup = bs4.BeautifulSoup(r.content, 'html.parser')
        # llamo funcion
        parsear_pagina(fullweb_soup, compras, keywords, web)   
    
    armar_excel(compras)
    
    
if __name__ == '__main__':
    keywords = ['viveres', 'desayunos', 'alimentacion', 'pan', 'lacteos',
                'pollo', 'fruta', 'frutas', 'verduras', 'verdura',
                'pastas', 'pasta', 'catering', 'raciones', 'almuerzos',
                'cenas', 'meriendas', 'sobrealimentacion', 'concesion',
                'hotel', 'comedor', 'automotor', 'camioneta', 
                'service', 'repuestos', 'reparaciones', 'tapiceria',
                'sillones', 'sillas']
    main(keywords)
    